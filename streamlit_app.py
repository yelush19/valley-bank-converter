import streamlit as st
import csv
import io
import re
import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import get_column_letter, quote_sheetname

# ===== PAGE CONFIG =====
st.set_page_config(page_title="ממיר תנועות בנק ואשראי", page_icon="🏦", layout="wide")

# ===== COA LOOKUP TABLES =====
PAYEM_COA_DEFAULT = {"LEVERAGED OUTBOUND|Brandlight Inc.":["502010","911000"],"Upwork|Brandlight Inc.":["502010","962006"],"Calendly|Brandlight Inc.":["502010","901009"],"Linkedin|Brandlight Inc.":["911004","502010"],"CONTENT ZEN|Brandlight Inc.":["502010","901009"],"PAYPAL *MADHAVMISTR|Brandlight Inc.":["502010","964009"],"TELMA SHLOMO|Brandlight AI LTD":["300001","921008"],"Medium|Brandlight Inc.":["502010","901009"],"OpenAI|Brandlight Inc.":["901009","502010"],"TAPLIO.COM/B|Brandlight Inc.":["502010","911004"],"UPS|Brandlight AI LTD":["300001","965001"],"PayEm Cash Back|Brandlight Inc.":["969001","502010"],"UPWORK * -842329291|Brandlight Inc.":["502010","962006"],"MISRAD HAPNIM|Brandlight Inc.":["502010","964009"],"AIRBNB * HM8RCRNBBT|Brandlight AI LTD":["300001","921008"],"AIRALO|Brandlight Inc.":["502010","921008"],"ISRAIR FLIGHT AND TOURISM|Brandlight Inc.":["502010","921008"],"DELAWARE CORP & TAX WE|Brandlight Inc.":["502010","964010"],"UPWORK * -844557510|Brandlight Inc.":["502010","962006"],"GT GET TAXI SYSTEMS LTD|Brandlight Inc.":["502010","965001"],"BITLY.COM|Brandlight Inc.":["502010","901009"],"Google|Brandlight Inc.":["502010","901009"],"UPWORK * -846744433|Brandlight Inc.":["502010","962006"],"Bolt|Brandlight Inc.":["502010","921008"],"BOLT.EU/O/2509151956|Brandlight Inc.":["502010","921008"],"BOLT.EU/O/2509151958|Brandlight Inc.":["502010","921008"],"PAIN FAUBOURG|Brandlight Inc.":["502010","963000"],"HORTENSE|Brandlight Inc.":["502010","964009"],"RESTAUR MARLOE|Brandlight Inc.":["502010","963000"],"BOLT.EU/O/2509170231|Brandlight Inc.":["502010","965001"],"BOLT.EU/O/2509170235|Brandlight Inc.":["502010","965001"],"Uber|Brandlight Inc.":["502010","965001"],"GRAN VIEW APARTMENTS|Brandlight Inc.":["502010","921008"],"Subway|Brandlight Inc.":["502010","963000"],"GranVia52|Brandlight Inc.":["502010","963000"],"MARKET TUDESCOS|Brandlight Inc.":["502010","963000"],"STARBUCKS GRAN VIA 46|Brandlight Inc.":["502010","963000"],"PAIPAI|Brandlight Inc.":["502010","963000"],"AEROP. ADOLFO SUAREZ MADR|Brandlight Inc.":["502010","963000"],"BARCELONA AIRPORTHOTEL|Brandlight Inc.":["502010","921008"],"BO I FRESC|Brandlight Inc.":["502010","963000"],"LA TORNA CATERINA|Brandlight Inc.":["502010","963000"],"CUINA SANTA CATERINA|Brandlight Inc.":["502010","963000"],"STARBUCKS COFFEE VIA LAIE|Brandlight Inc.":["502010","963000"],"TAXI LIC.6475|Brandlight Inc.":["502010","965001"],"626 - PC AER BCN T1 AIRE|Brandlight Inc.":["502010","963000"],"Immfly SL-Vta a Bordo|Brandlight Inc.":["502010","963000"],"SQ *DELECTICA|Brandlight Inc.":["502010","963000"],"SQ *CAFE #65095 UNION|Brandlight Inc.":["502010","963000"],"Whole Foods Market|Brandlight Inc.":["502010","963000"],"Lyft|Brandlight Inc.":["502010","965001"],"FRESH DIRECT|Brandlight Inc.":["502010","963000"],"PRET A MANGER US0037|Brandlight Inc.":["502010","963000"],"MTA|Brandlight Inc.":["502010","965001"],"Walgreens|Brandlight Inc.":["502010","964009"],"PARKER QUINN RESTAURANT|Brandlight Inc.":["502010","963000"],"UPWORK * -848944854|Brandlight Inc.":["502010","962006"],"Starbucks|Brandlight Inc.":["502010","963000"],"THE ATHENA PROJECT|Brandlight Inc.":["502010","911004"],"SPIKES STUDIO|Brandlight Inc.":["502010","400002"],"STARBUCKS STORE 70231|Brandlight Inc.":["502010","963000"],"Pershing|Brandlight Inc.":["502010","965001"],"VESTA  *VODAFONE TOPUP|Brandlight Inc.":["502010","964009"],"SQ *THINK NOMAD LLC|Brandlight Inc.":["502010","963000"],"Trader Joe's|Brandlight Inc.":["502010","963000"],"Leveraged Outbound Ltd|Brandlight Inc.":["502010","911004"],"PARK EAST CLEANERS|Brandlight Inc.":["502010","964009"],"SQ *KONA COFFEE ROASTE|Brandlight Inc.":["502010","963000"],"Soho House|Brandlight Inc.":["502010","911004"],"WOLFGANG STEAKHOUSE PA|Brandlight Inc.":["502010","963000"],"THE CLOCK COFFEE SHO|Brandlight Inc.":["502010","963000"],"CVS Pharmacy|Brandlight Inc.":["502010","964009"],"NYC Ferry|Brandlight Inc.":["502010","965001"],"OTTOMANELLI LIC|Brandlight Inc.":["502010","963000"],"TST* THE STANDARD GRILL|Brandlight Inc.":["502010","963000"],"TST* BIERGARTEN|Brandlight Inc.":["502010","963000"],"GREENWICH GOURMET MARK|Brandlight Inc.":["502010","963000"],"TST* GREGORY'S COFFEE GC0|Brandlight Inc.":["502010","963000"],"D'agostino|Brandlight Inc.":["502010","963000"],"SP FELT RIGHT LLC|Brandlight Inc.":["502010","911004"],"TENDER GREENS MDR|Brandlight Inc.":["502010","963000"],"CLAUDE.AI SUBSCRIPTION|Brandlight Inc.":["502010","901009"],"WWW.PODCASTLE.AI|Brandlight Inc.":["502010","911004"],"Uber Eats|Brandlight Inc.":["502010","963000"],"NYC TAXI 1246|Brandlight Inc.":["502010","965001"],"RIVERSIDEFM, INC.|Brandlight Inc.":["964009","502010"],"THE SMITH - LINCOLN|Brandlight Inc.":["502010","963000"],"Chipotle Mexican Grill|Brandlight Inc.":["502010","963000"],"TST* GREGORYS COFFEE - GC|Brandlight Inc.":["502010","963000"],"TST* JOE COFFEE - BRYANT|Brandlight Inc.":["502010","963000"],"Burlington Stores|Brandlight Inc.":["502010","964009"],"CINICO ON MADISON|Brandlight Inc.":["502010","963000"],"SQ *THE COFFEE HAUS|Brandlight Inc.":["502010","963000"],"BKG*Hotel at Booking.c|Brandlight Inc.":["502010","921008"],"TST* MAMAN - KING|Brandlight Inc.":["502010","963000"],"British Airways|Brandlight Inc.":["502010","921008"],"LOBSTER PLACE|Brandlight Inc.":["502010","963000"],"Curb Mobility|Brandlight Inc.":["502010","965001"],"VIRGIN NYC - AITANA|Brandlight Inc.":["502010","963000"],"Dropbox|Brandlight Inc.":["901009","502010"],"MAMAN|Brandlight Inc.":["502010","963000"],"Pret A Manger|Brandlight Inc.":["502010","963000"],"Amazon|Brandlight Inc.":["964009","502010"],"LA PECORA BIANCA 8|Brandlight Inc.":["502010","963000"],"Sainsbury's|Brandlight Inc.":["502010","963000"],"GOOGLE *ADS9696437247|Brandlight Inc.":["502010","901009"],"Sweetgreen|Brandlight Inc.":["502010","963000"],"NOA Cafe|Brandlight Inc.":["502010","963000"],"Modern Bread  Bagel|Brandlight Inc.":["502010","963000"],"MODERN B&B MIDTOWN N|Brandlight Inc.":["502010","963000"],"UPWORK * -855585411|Brandlight Inc.":["502010","962006"],"Apple|Brandlight Inc.":["502010","400002"],"GLAZE - UNION SQUARE|Brandlight Inc.":["502010","963000"],"MODERN BREAD & BAGEL|Brandlight Inc.":["502010","963000"],"THE PERFECT GIFT INC|Brandlight Inc.":["502010","964009"],"TST* JEAN'S NYC|Brandlight Inc.":["502010","963000"],"United|Brandlight Inc.":["502010","921008"],"MAMA MEZZE|Brandlight Inc.":["502010","963000"],"TST* FASANO RESTAURANT|Brandlight Inc.":["502010","963000"],"TST* SIP AND CO 1|Brandlight Inc.":["502010","963000"],"KLM Airlines|Brandlight Inc.":["502010","921008"],"SUPER PHARM VR|Brandlight AI LTD":["300001","964009"],"PROFOUND AI|Brandlight Inc.":["502010","911004"],"Playa Bowls|Brandlight Inc.":["502010","963000"],"Shipi smartphone|Brandlight AI LTD":["300001","400002"],"WWW.MAILERLITE.COM|Brandlight Inc.":["911004","502010"],"CONTENT ZEN|Brandlight AI LTD":["300001","901009"],"Apple|Brandlight AI LTD":["502010","400002"],"RUBIROSA - MULBERRY|Brandlight Inc.":["502010","963000"],"TELMA SHLOMO|Brandlight Inc.":["502010","921008"],"LYRIC HOTEL|Brandlight Inc.":["502010","921008"],"COMMON HOURS|Brandlight Inc.":["502010","963000"],"SQ *DON CAFE|Brandlight Inc.":["502010","963000"],"Shake Shack|Brandlight Inc.":["502010","963000"],"IDIGITAL STORE ITD|Brandlight AI LTD":["300001","400002"],"Adobe|Brandlight Inc.":["502010","901009"],"DD/BR #348531|Brandlight Inc.":["502010","963000"],"BLUEGROUND US, INC.|Brandlight Inc.":["502010","921008"],"DD-BR #358650|Brandlight Inc.":["963000","502010"],"GOOGLE*ADS9696437247|Brandlight Inc.":["901009","502010"],"Delta Airlines|Brandlight Inc.":["921008","502010"],"Notion|Brandlight Inc.":["901009","502010"],"BRIGHTDATA|Brandlight AI LTD":["901009","300001"],"OpenAI|Brandlight AI LTD":["901009","300001"],"TOWNEPLACE STS SLT LK CTY|Brandlight Inc.":["921008","502010"],"Serpapi|Brandlight AI LTD":["901009","300001"],"HUGO COFFEE ROASTERS SLC|Brandlight Inc.":["963000","502010"],"SLC Vessel Kitchen|Brandlight Inc.":["963000","502010"],"CAPCUT|Brandlight AI LTD":["901009","300001"],"HARMONS - BANGERTER 16|Brandlight Inc.":["964009","502010"],"ANTHROPIC  CLAUDE TEAM|Brandlight AI LTD":["901009","300001"],"protein Bar & Kitchen|Brandlight Inc.":["963000","502010"],"SLC iStore Exp/Relay|Brandlight Inc.":["400002","502010"],"The Coffee & Tea Exchange|Brandlight Inc.":["963000","502010"],"OHM - GOURMANDISE|Brandlight Inc.":["963000","502010"],"PayMe|Brandlight AI LTD":["962006","300001"],"CURB NYC TAXI|Brandlight Inc.":["965001","502010"],"ANTHROPIC|Brandlight AI LTD":["901009","300001"],"easy coffee|Brandlight Inc.":["963000","502010"],"PAYPAL *ALPHAQUANTU|Brandlight AI LTD":["964009","300001"],"Vcorp Services|Brandlight Inc.":["962006","502010"],"SUZANA|Brandlight Inc.":["963000","502010"],"SITE-SHOT.COM|Brandlight AI LTD":["901009","300001"],"Figma|Brandlight AI LTD":["901009","300001"],"PADDLE.NET* N8N CLOUD1|Brandlight AI LTD":["901009","300001"],"2M CARBURANTI|Brandlight Inc.":["965001","502010"],"SCIROCCO|Brandlight Inc.":["963000","502010"],"Delicia Imb Fiumicino T1|Brandlight Inc.":["963000","502010"],"ONLINE PAYMENT|Brandlight AI LTD":["962006","300001"],"BOLT.EUO2511121449|Brandlight Inc.":["965001","502010"],"Webflow|Brandlight AI LTD":["901009","300001"],"Barmhartig Koffie|Brandlight Inc.":["963000","502010"],"Webflow|Brandlight Inc.":["901009","502010"],"Jumbo O.handelskade|Brandlight Inc.":["963000","502010"],"The Hoxton|Brandlight Inc.":["921008","502010"],"Ahrefs|Brandlight Inc.":["901009","502010"],"BOLT.EUO2511132019|Brandlight Inc.":["965001","502010"],"LS Madre|Brandlight Inc.":["963000","502010"],"BOLT.EUO2511132142|Brandlight Inc.":["965001","502010"],"SCRUNCH AI|Brandlight Inc.":["901009","502010"],"BKG*HOTEL AT BOOKING.C|Brandlight Inc.":["921008","502010"],"NACHAT|Brandlight Inc.":["963000","502010"],"BOLT.EUO2511160920|Brandlight Inc.":["965001","502010"],"HMS Host International|Brandlight Inc.":["963000","502010"],"Avolta AAMS Loaf 2203 009|Brandlight Inc.":["963000","502010"],"Viasat|Brandlight Inc.":["901009","502010"],"GRAMMARLY CO*IJBTT8J|Brandlight Inc.":["901009","502010"],"HOTEL AURA TIMES SQUARE|Brandlight Inc.":["921008","502010"],"FS *dataforseo|Brandlight AI LTD":["901009","300001"],"CAFFE VITA - LOWER E|Brandlight Inc.":["963000","502010"],"TST* DUDLEYS|Brandlight Inc.":["963000","502010"],"PEEC AI|Brandlight AI LTD":["901009","300001"],"TRINITY PLACE RESTAURA|Brandlight Inc.":["963000","502010"],"Deliveroo|Brandlight Inc.":["963000","502010"],"TST* FRIEDMAN'S 31 STREET|Brandlight Inc.":["963000","502010"],"LA BOMBE|Brandlight Inc.":["963000","502010"],"Twilio|Brandlight AI LTD":["901009","300001"],"LEMSQZY* SCREENSTUDIO|Brandlight AI LTD":["901009","300001"],"Cursor|Brandlight AI LTD":["901009","300001"],"Linkedin|Brandlight AI LTD":["911004","300001"],"FORBES BUSINESS COUNCIL|Brandlight Inc.":["964009","502010"],"AIRBNB * HMC3JYCAHJ|Brandlight AI LTD":["921008","502010"],"CURSOR USAGE MID  NOV|Brandlight AI LTD":["901009","300001"],"DEEPL* SUB 198XSOV0PVF|Brandlight AI LTD":["964009","300001"],"Notion|Brandlight AI LTD":["901009","300001"],"BASE44|Brandlight AI LTD":["901009","300001"],"SUPER|Brandlight AI LTD":["963000","300001"],"Docusign|Brandlight AI LTD":["901009","300001"],"PAYPAL *EVENTHANDLE EVENT|Brandlight AI LTD":["911004","300001"],"HERCOLES LOCKS|Brandlight AI LTD":["400002","300001"],"Amtrak|Brandlight Inc.":["965001","502010"],"GALACI LTD|Brandlight AI LTD":["911004","300001"],"Ashby|Brandlight AI LTD":["901009","300001"],"NUTT LABS* NOTION VIP|Brandlight Inc.":["901009","502010"],"ISTORE|Brandlight AI LTD":["400002","300001"],"MEETALFRED.COM|Brandlight AI LTD":["901009","300001"],"Github|Brandlight AI LTD":["901009","300001"],"Zoho|Brandlight AI LTD":["901009","300001"],"ZOHO* ZOHO-BOOKS|Brandlight AI LTD":["901009","300001"],"BUY ME|Brandlight AI LTD":["963000","300001"],"CURSOR USAGE  NOV|Brandlight AI LTD":["901009","300001"],"MAX 10|Brandlight AI LTD":["963000","300001"],"ZIG - ZAG|Brandlight AI LTD":["965001","300001"],"SEARCHABLE.COM|Brandlight Inc.":["901009","502010"],"WWW.PERPLEXITY.AI|Brandlight AI LTD":["901009","300001"],"SCRUNCH|Brandlight Inc.":["901009","502010"],"EXPEDIA 72069208160332|Brandlight Inc.":["921008","502010"],"CURSOR USAGE MID  DEC|Brandlight AI LTD":["901009","300001"],"Tatte Bakery Cafe|Brandlight Inc.":["963000","502010"],"OLO*123 honeygrow Seap|Brandlight Inc.":["963000","502010"],"Flight Club|Brandlight Inc.":["921008","502010"],"TST* MARCELINO'S BOUTIQUE|Brandlight Inc.":["963000","502010"],"TST* LOLITA COCINA & TEQU|Brandlight Inc.":["963000","502010"],"RESHUT HADOAR MECHIR|Brandlight AI LTD":["965001","300001"],"SUPERYUDA|Brandlight AI LTD":["963000","300001"],"I LOVE CUPCAKES|Brandlight AI LTD":["963000","300001"],"Sentry|Brandlight AI LTD":["901009","300001"],"SOPER KLIL|Brandlight AI LTD":["967002","300001"],"PayEm Cash Back|Brandlight AI LTD":["969001","300001"]}

VALLEY_COA_DEFAULT = {"FISHER INVESTMEN INV- RMR*IK*INV-\\|Revenue":["100000","700000"],"NATIONAL EMPLOYE INVOICE H|Employee Salaries & Benefits":["540001","100000"],"Deel|Employee Salaries & Benefits":["500009","100000"],"The Yard 2 Lower SIGONFILE|Office Rent":["964000","100000"],"BRANDLIGHT LTD LEVERAGED|Marketing & Advertising":["911003","100000"],"NSKNOX TECHNOLOG PAYMENTS|Bank Fees & interest":["100000","990001"],"PRIME ONLINE LIMITED BRANDLIGHT INC. INV|Revenue":["100000","700000"],"PHILIP STEIN AND ASSOCIATES LTD PRO FORM|Professional Services":["962006","100000"],"NSKNOX TECHNOLOGIES|Bank Fees & interest":["990001","100000"],"PAYEM CSMRS PRGM FUNDING BRANDLIGHT AI|Credit/Debit banks/Accounts":["502010","100000"],"CENSUSWIDE LTD BRANDLIGHT REMAINING INV.|Marketing & Advertising":["911003","100000"],"BRANDLIGHT AI LTD|Credit/Debit banks/Accounts":["100000","300001"],"SAS CADOA BRANDLIGHT CAD2442|Technology":["901005","100000"],"DEAL# 1449182|Marketing & Advertising":["911003","100000"],"MMS AS PAYCO6366 EDI PAYMNT REF*TN** Var|Revenue":["100000","700000"],"IRS USER FEE PAYMENT|Authorities":["990002","100000"],"ANDREW BLOOM CONSULTING LLC BRANDLIGHT I|Vendors":["500001","100000"],"CONNEXUS LIMITED BRANDLIGHT INVOICE NO34|Marketing & Advertising":["911003","100000"],"ANALYSIS ACTIVITY FOR 10/25|Bank Fees & interest":["990001","100000"],"ATASH LAW P.C. BRANDLIGHT INV 2281|Professional Services":["962006","100000"],"HUBSPOT|Marketing & Advertising":["911003","100000"],"LOGMEIN 11.19.2025 RMR*IK*IN\\|Revenue":["100000","700000"],"LEO BURNETT LIMITED /BNF//CHPREF/|Revenue":["100000","700000"],"Teneo Strategy L USSC 11.21 00004599/|Revenue":["100000","700000"],"ARNON|Professional Services":["962006","100000"],"BRANDLIGHT LTD PUBLIC REL|Marketing & Advertising":["911003","100000"],"ISCAR LTD INVOICES|Revenue":["100000","700000"],"KIMBERLY CL6933 EDI PAYMNT|Revenue":["100000","700000"],"1/SAAR TECHNOLOGIES )Z.H( LTD /BNF/BOOK/|Revenue":["100000","700000"],"WISE US INC From Inmarkets Limited Via W|Revenue":["100000","700000"],"SOFWAVE MEDICAL LTD INVOICE NO. 034 DATE|Marketing & Advertising":["100000","700000"],"ANDREW BLOOM CONSULTING LLC BRANDLIGHT I|Marketing & Advertising":["911003","100000"],"SOFWAVE MEDICAL LTD INVOICE NO. 034 DATE|Revenue":["100000","700000"],"XVERUM LLC BRANDLIGHT|Professional Services":["962006","100000"],"DAVID SHIELD - PASSPORTCARD LTD INV- PAY|Revenue":["100000","700000"],"1/ZYG EDGE LTD INV- /BNF/BOOK/JS|Revenue":["100000","700000"],"ANALYSIS ACTIVITY FOR 11/25|Bank Fees & interest":["990001","100000"],"TENEO STRATEGY LIMITED /ROC/BRANDLIGHT I|Revenue":["100000","700000"],"DATA AXLE PJP ACH 12|Revenue":["100000","700000"],"PELION VENTURES VIII-A|Investment":["100000","600015"],"PELION VENTURES VIII-C|Investment":["100000","600015"],"PELION VENTURES VIII - ENTREPRENEU+ PELI|Investment":["100000","600016"],"PELION VENTURES VIII FINANCIAL PELION VE|Investment":["100000","600017"],"PELION VENTURES VIII|Investment":["100000","600015"],"XVERUM LLC BRANDLIGHT|Recruitment":["962003","100000"],"HARTFORD FIRE IN TRADE EXC|Revenue":["100000","700000"],"G20 VENTURES IV LP BRANDLIGHT INC SAFE|Investment":["100000","600006"],"Deposit|Revenue":["100000","700000"],"LLORENTE CUENCA MADRID SL /ROC/AFJQ18/UR|Revenue":["100000","700000"],"CARDUMEN DEEPTECH II ILS SCR-PYME ABA/RO|Investment":["100000","600018"],"CARDUMEN DEEPTECH II A SCR-PYME S PLEASE|Investment":["100000","600019"],"CARDUMEN DEEPTECH FUND II FCRE PLEASE PA|Investment":["100000","600020"],"MAYER BROWN LLP 25802709 INVESTMENT INBR|Professional Services":["962006","100000"],"GOOGLE ACCTVERIFY US004A0Q|Bank Fees & interest":["990001","100000"]}

VALLEY_CAT_COA = {"Revenue":["100000","700000"],"Employee Salaries & Benefits":["540001","100000"],"Office Rent":["964000","100000"],"Marketing & Advertising":["911003","100000"],"Bank Fees & interest":["990001","100000"],"Professional Services":["962006","100000"],"Credit/Debit banks/Accounts":["502010","100000"],"Technology":["901005","100000"],"Authorities":["990002","100000"],"Vendors":["500001","100000"],"Investment":["100000","600015"],"Recruitment":["962003","100000"]}

LTD_COA_DEFAULT = {"TELMA SHLOMO":["104000","921008"],"UPS":["104000","965001"],"AIRBNB * HM8RCRNBBT":["104000","921008"],"SUPER PHARM VR":["104000","964009"],"Shipi smartphone":["104000","400002"],"CONTENT ZEN":["104000","901009"],"Apple":["104000","400002"],"IDIGITAL STORE ITD":["104000","400002"],"BRIGHTDATA":["104000","901009"],"OpenAI":["104000","901009"],"Serpapi":["104000","901009"],"CAPCUT":["104000","901009"],"ANTHROPIC  CLAUDE TEAM":["104000","901009"],"PayMe":["104000","962006"],"ANTHROPIC":["104000","901009"],"PAYPAL *ALPHAQUANTU":["104000","964009"],"SITE-SHOT.COM":["104000","901009"],"Figma":["104000","901009"],"PADDLE.NET* N8N CLOUD1":["104000","901009"],"ONLINE PAYMENT":["104000","962006"],"Webflow":["104000","901009"],"FS *dataforseo":["104000","901009"],"PEEC AI":["104000","901009"],"Twilio":["104000","901009"],"LEMSQZY* SCREENSTUDIO":["104000","901009"],"Cursor":["104000","901009"],"Linkedin":["104000","911004"],"AIRBNB * HMC3JYCAHJ":["104000","921008"],"CURSOR USAGE MID  NOV":["104000","901009"],"DEEPL* SUB 198XSOV0PVF":["104000","964009"],"Notion":["104000","901009"],"BASE44":["104000","901009"],"SUPER":["104000","963000"],"Docusign":["104000","901009"],"PAYPAL *EVENTHANDLE EVENT":["104000","911004"],"HERCOLES LOCKS":["104000","400002"],"GALACI LTD":["104000","911004"],"Ashby":["104000","901009"],"ISTORE":["104000","400002"],"MEETALFRED.COM":["104000","901009"],"Github":["104000","901009"],"Zoho":["104000","901009"],"ZOHO* ZOHO-BOOKS":["104000","901009"],"BUY ME":["104000","963000"],"CURSOR USAGE  NOV":["104000","901009"],"MAX 10":["104000","963000"],"ZIG - ZAG":["104000","965001"],"WWW.PERPLEXITY.AI":["104000","901009"],"CURSOR USAGE MID  DEC":["104000","901009"],"RESHUT HADOAR MECHIR":["104000","965001"],"SUPERYUDA":["104000","963000"],"I LOVE CUPCAKES":["104000","963000"],"Sentry":["104000","901009"],"SOPER KLIL":["104000","967002"]}


# ===== HELPER FUNCTIONS =====

def detect_file_type(rows):
    """Auto-detect CSV type: valley or payem."""
    if not rows or len(rows) < 2:
        return "unknown"
    row0 = [str(c or "") for c in rows[0]]
    row1 = [str(c or "") for c in rows[1]]
    if any("Valley" in c for c in row0):
        return "valley"
    if any("PayEm" in c for c in row0):
        return "payem"
    if row1 and row1[0].startswith("Bank"):
        return "valley"
    if len(row1) >= 3 and row1[0] == "Date" and row1[1] == "Time" and row1[2] == "Status":
        return "payem"
    return "unknown"


def clean_description(bai_desc, detail):
    """Clean Valley transaction description to max 40 chars."""
    c = detail or ""
    c = re.sub(r"^20\d{11,}\s*", "", c)
    c = re.sub(r"\d{10,}", "", c)
    c = re.sub(r"\b\d{6}\b", "", c)
    c = re.sub(r"ST-?[A-Z0-9]{10,}", "", c)
    c = re.sub(r"\d{2}/\d{2}/\d{2}", "", c)
    c = re.sub(r"\s+", " ", c).strip()
    if len(c) > 40:
        c = c[:40].strip()
    if not c or len(c) < 3:
        return bai_desc
    return c


def parse_amount(s):
    """Parse Valley amount like $36,000.00 or ($50,000.00)."""
    c = re.sub(r"[\$,\s]", "", str(s or ""))
    c = re.sub(r"[()]", "", c)
    try:
        return abs(float(c))
    except (ValueError, TypeError):
        return 0.0


def parse_net_amount(s):
    """Parse PayEm net amount like ($1,703) -> -1703."""
    c = re.sub(r"[\$,\s]", "", str(s or ""))
    neg = "(" in c
    c = re.sub(r"[()]", "", c)
    try:
        v = float(c)
    except (ValueError, TypeError):
        v = 0.0
    if neg:
        v = -v
    return v


def parse_valley_date(s):
    """Parse Valley date MM/DD/YYYY -> datetime."""
    parts = str(s or "").split("/")
    if len(parts) != 3:
        return None
    try:
        m, d, y = int(parts[0]), int(parts[1]), int(parts[2])
        return datetime(y, m, d)
    except (ValueError, TypeError):
        return None


def parse_payem_date(s):
    """Parse PayEm date YYYY-MM-DD -> datetime."""
    parts = str(s or "").split("-")
    if len(parts) != 3:
        return None
    try:
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
        return datetime(y, m, d)
    except (ValueError, TypeError):
        return None


def format_date(dt):
    """Format datetime as DD/MM/YYYY."""
    return f"{dt.day:02d}/{dt.month:02d}/{dt.year}"


def generate_valley_key(date_str, amount, ref, bai_desc, detail):
    """Generate unique key for Valley transaction dedup."""
    r = str(ref or "").strip()
    if r and r != "0":
        return f"{date_str}|{amount}|{r}"
    return f"{date_str}|{amount}|{bai_desc}|{(detail or '')[:20]}"


# ===== PARSERS =====

def parse_valley(rows, coa_lookup, cat_coa):
    """Parse Valley Bank CSV rows into transaction dicts."""
    results = []
    for i in range(2, len(rows)):
        row = rows[i]
        if not row or len(row) < 12:
            continue
        bank_aba = str(row[0] or "")
        if bank_aba.startswith("Bank"):
            continue
        date_str = str(row[5] or "")
        type_ind = str(row[6] or "").upper()
        amount_str = str(row[7] or "")
        cust_ref = str(row[8] or "")
        bai_desc = str(row[10] or "").upper()
        detail = str(row[11] or "")
        category = str(row[20] or "") if len(row) > 20 else ""

        is_valid = (
            "ACH" in bai_desc or "WIRE" in bai_desc or "DEPOSIT" in bai_desc
            or "FEE" in bai_desc or type_ind in ("DEBIT", "CREDIT")
        )
        if not is_valid or "/" not in date_str or not amount_str:
            continue

        dt = parse_valley_date(date_str)
        if not dt:
            continue
        amount = parse_amount(amount_str)

        desc = clean_description(bai_desc, detail)
        key = generate_valley_key(date_str, amount, cust_ref, bai_desc, detail)

        coa_key = f"{desc}|{category}"
        coa_match = coa_lookup.get(coa_key)
        cat_match = cat_coa.get(category)
        coa_debit = coa_match[0] if coa_match else (cat_match[0] if cat_match else "")
        coa_credit = coa_match[1] if coa_match else (cat_match[1] if cat_match else "")

        results.append({
            "key": key, "date": dt, "date_formatted": format_date(dt),
            "description": desc, "amount": amount, "category": category,
            "coa_debit": coa_debit, "coa_credit": coa_credit,
        })

    results.sort(key=lambda t: t["date"])
    return results


def parse_payem(rows, coa_lookup, ltd_coa):
    """Parse PayEm CSV rows into transaction dicts."""
    results = []
    for i in range(2, len(rows)):
        row = rows[i]
        if not row or len(row) < 42:
            continue
        status = str(row[2] or "").upper()
        if status != "CLEARED":
            continue
        date_str = str(row[0] or "")
        dt = parse_payem_date(date_str)
        if not dt:
            continue

        merchant = str(row[28] or "").strip()
        txn_id = str(row[41] or "").strip()
        card4 = str(row[35] or "").strip()
        subsidiary = str(row[9] or "").strip()
        net_str = str(row[53] or "") if len(row) > 53 else str(row[5] or "")
        net_amount = parse_net_amount(net_str)
        abs_amount = abs(net_amount)
        is_credit = net_amount > 0
        credit_amount = abs_amount if is_credit else None
        debit_amount = abs_amount if not is_credit else None

        coa_key = f"{merchant}|{subsidiary}"
        coa_match = coa_lookup.get(coa_key)
        coa_credit = coa_match[0] if coa_match else ""
        coa_debit = coa_match[1] if coa_match else ""

        ltd_match = ltd_coa.get(merchant)
        ltd_coa_credit = ltd_match[0] if ltd_match else ""
        ltd_coa_debit = ltd_match[1] if ltd_match else ""

        try:
            ref1_num = int(txn_id)
        except (ValueError, TypeError):
            ref1_num = txn_id

        results.append({
            "key": txn_id, "date": dt, "date_formatted": format_date(dt),
            "description": merchant, "net_amount": net_amount,
            "abs_amount": abs_amount, "credit_amount": credit_amount,
            "debit_amount": debit_amount, "ref1": ref1_num, "ref2": card4,
            "coa_credit": coa_credit, "coa_debit": coa_debit,
            "ltd_coa_credit": ltd_coa_credit, "ltd_coa_debit": ltd_coa_debit,
            "subsidiary": subsidiary,
        })

    results.sort(key=lambda t: t["date"])
    return results


# ===== DEDUP =====

def analyze_duplicates(file_type, transactions, history):
    """Split transactions into new and duplicate."""
    known_keys = history.get(file_type, {}).get("keys", {})
    new_txns = [t for t in transactions if t["key"] not in known_keys]
    dupes = [t for t in transactions if t["key"] in known_keys]
    return new_txns, dupes


# ===== EXCEL BUILDER =====

def _format_text_col(ws, col, min_row, max_row):
    """Format a column as text (@ format) with string values."""
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=col, max_col=col):
        for cell in row:
            if cell.value is not None:
                cell.value = str(cell.value)
                cell.number_format = '@'


def _add_named_range(wb, name, ws_title, min_col, min_row, max_col, max_row):
    """Add a named range (not table) to workbook."""
    ref = f"{quote_sheetname(ws_title)}!${get_column_letter(min_col)}${min_row}:${get_column_letter(max_col)}${max_row}"
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)


def build_valley_excel(data):
    """Build Valley Bank Excel with VALLEYTRANS sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "VALLEYTRANS"
    headers = ["תאריך", "תיאור", "חובה", "זכות", "קטגוריה", 'ח"ן חובה', 'ח"ן זכות']
    ws.append(headers)
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")

    for t in data:
        ws.append([
            t["date_formatted"], t["description"], t["amount"], t["amount"],
            t["category"], t["coa_debit"], t["coa_credit"]
        ])

    # Format number columns
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
        for cell in row:
            cell.number_format = '#,##0.00'

    # Format CoA columns as text (F=coa_debit, G=coa_credit)
    _format_text_col(ws, 6, 2, ws.max_row)  # F = ח"ן חובה
    _format_text_col(ws, 7, 2, ws.max_row)  # G = ח"ן זכות

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12

    # Named range for data area (excluding header)
    if len(data) > 0:
        _add_named_range(wb, "VALLEYTRANS", "VALLEYTRANS", 1, 2, 7, ws.max_row)

    return wb


def build_payem_excel(data):
    """Build PayEm Excel with 4 sheets. Returns (workbook, integrity_checks)."""
    wb = Workbook()
    wb.remove(wb.active)

    inc_data = [t for t in data if "Inc" in t["subsidiary"]]
    ltd_data = [t for t in data if "LTD" in t["subsidiary"]]

    _add_payem_sheet(wb, data, "PAYEMDATA", "all")
    _add_payem_sheet(wb, inc_data, "פקודה INC", "inc")
    _add_payem_sheet(wb, ltd_data, "רישום LTD", "ltd")
    _add_payem_sheet(wb, ltd_data, "נגדית לINC של LTD", "negdit")

    # ===== INTEGRITY CHECKS =====
    # Check 1: INC + LTD = PAYEMDATA (row count)
    count_ok = len(inc_data) + len(ltd_data) == len(data)

    # Check 2: sum debit in DATA = sum debit in INC + LTD
    data_debit = sum(t["abs_amount"] for t in data if t["net_amount"] <= 0)
    data_credit = sum(t["abs_amount"] for t in data if t["net_amount"] > 0)
    inc_debit = sum(t["abs_amount"] for t in inc_data if t["net_amount"] <= 0)
    inc_credit = sum(t["abs_amount"] for t in inc_data if t["net_amount"] > 0)
    ltd_debit = sum(t["abs_amount"] for t in ltd_data if t["net_amount"] <= 0)
    ltd_credit = sum(t["abs_amount"] for t in ltd_data if t["net_amount"] > 0)

    debit_ok = abs(data_debit - (inc_debit + ltd_debit)) < 0.01
    credit_ok = abs(data_credit - (inc_credit + ltd_credit)) < 0.01

    checks = {
        "count_ok": count_ok,
        "count_data": len(data), "count_inc": len(inc_data), "count_ltd": len(ltd_data),
        "debit_ok": debit_ok,
        "data_debit": data_debit, "inc_debit": inc_debit, "ltd_debit": ltd_debit,
        "credit_ok": credit_ok,
        "data_credit": data_credit, "inc_credit": inc_credit, "ltd_credit": ltd_credit,
    }

    return wb, checks


def _add_payem_sheet(wb, data, sheet_name, mode):
    """Add a PayEm sheet to workbook.
    mode: 'all' = PAYEMDATA (raw data with Transaction amount)
          'inc' = פקודה INC
          'ltd' = רישום LTD
          'negdit' = נגדית לINC של LTD
    """
    ws = wb.create_sheet(title=sheet_name)

    if mode == "all":
        # PAYEMDATA = raw data, includes Transaction amount
        headers = ["תיאור", "Transaction amount", "זכות מטח", "חובה מטח",
                   "אסמכתא 1", "אסמכתא 2", 'ח"ן זכות', 'ח"ן חובה', "תאריך", "שיוך"]
    else:
        # Pekuda sheets: no Transaction amount column
        headers = ["תיאור", "זכות מטח", "חובה מטח",
                   "אסמכתא 1", "אסמכתא 2", 'ח"ן זכות', 'ח"ן חובה', "תאריך", "שיוך"]

    ws.append(headers)

    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="right")

    for t in data:
        is_credit = t["net_amount"] > 0  # positive = credit/refund

        # Determine CoA based on mode
        if mode == "ltd":
            base_coa_cr = t["ltd_coa_credit"] or ""
            base_coa_dr = t["ltd_coa_debit"] or ""
        elif mode == "negdit":
            base_coa_cr = "502010"
            base_coa_dr = "300001"
        elif mode == "inc":
            base_coa_cr = t["coa_credit"]
            base_coa_dr = t["coa_debit"]
        else:
            # PAYEMDATA - original CoA, no flip
            base_coa_cr = t["coa_credit"]
            base_coa_dr = t["coa_debit"]

        # For pekuda sheets (inc, ltd, negdit): flip CoA on credits
        if mode != "all" and is_credit:
            coa_cr = base_coa_dr  # flip!
            coa_dr = base_coa_cr  # flip!
        else:
            coa_cr = base_coa_cr
            coa_dr = base_coa_dr

        # Amount columns: always both = abs amount (double-entry)
        abs_amt = t["abs_amount"]

        if mode == "all":
            # PAYEMDATA: split credit/debit + include Transaction amount
            credit_val = t["credit_amount"]
            debit_val = t["debit_amount"]
            ws.append([
                t["description"], t["net_amount"], credit_val, debit_val,
                t["ref1"], t["ref2"], coa_cr, coa_dr, t["date"], t["subsidiary"]
            ])
        else:
            # Pekuda sheets: both columns = abs amount, no Transaction amount
            ws.append([
                t["description"], abs_amt, abs_amt,
                t["ref1"], t["ref2"], coa_cr, coa_dr, t["date"], t["subsidiary"]
            ])

    # Format columns based on mode
    if mode == "all":
        # PAYEMDATA: A=desc, B=txn_amt, C=credit, D=debit, E=ref1, F=ref2, G=coa_cr, H=coa_dr, I=date, J=sub
        num_cols = (2, 4)   # B,C,D = numbers
        ref1_col = 5        # E
        ref2_col = 6        # F
        coa_cr_col = 7      # G
        coa_dr_col = 8      # H
        date_col = 9        # I
        total_cols = 10
        col_widths = {"A": 30, "B": 16, "C": 14, "D": 14, "E": 14, "F": 10, "G": 12, "H": 12, "I": 12, "J": 22}
    else:
        # Pekuda: A=desc, B=credit, C=debit, D=ref1, E=ref2, F=coa_cr, G=coa_dr, H=date, I=sub
        num_cols = (2, 3)   # B,C = numbers
        ref1_col = 4        # D
        ref2_col = 5        # E
        coa_cr_col = 6      # F
        coa_dr_col = 7      # G
        date_col = 8        # H
        total_cols = 9
        col_widths = {"A": 30, "B": 14, "C": 14, "D": 14, "E": 10, "F": 12, "G": 12, "H": 12, "I": 22}

    for row in ws.iter_rows(min_row=2, min_col=num_cols[0], max_col=num_cols[1]):
        for cell in row:
            if cell.value is not None:
                cell.number_format = '#,##0.00'

    for row in ws.iter_rows(min_row=2, min_col=date_col, max_col=date_col):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD/MM/YYYY'

    # Format ref1, ref2, CoA columns as text
    _format_text_col(ws, ref1_col, 2, ws.max_row)
    _format_text_col(ws, ref2_col, 2, ws.max_row)
    _format_text_col(ws, coa_cr_col, 2, ws.max_row)
    _format_text_col(ws, coa_dr_col, 2, ws.max_row)

    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # Named range for data area (excluding header)
    if len(data) > 0:
        # Sanitize name: replace spaces and Hebrew with safe chars
        safe_name = sheet_name.replace(" ", "_").replace('"', '')
        # Named ranges must start with letter/underscore, only ASCII allowed
        name_map = {"PAYEMDATA": "PAYEMDATA", "פקודה_INC": "PEKUDA_INC",
                     "רישום_LTD": "RISHUM_LTD", "נגדית_לINC_של_LTD": "NEGDIT_LTD"}
        range_name = name_map.get(safe_name, safe_name)
        _add_named_range(wb, range_name, sheet_name, 1, 2, total_cols, ws.max_row)


def workbook_to_bytes(wb):
    """Convert openpyxl Workbook to bytes for download."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ===== SESSION STATE INIT =====

def init_session():
    if "history" not in st.session_state:
        st.session_state["history"] = {"valley": {"keys": {}, "count": 0}, "payem": {"keys": {}, "count": 0}}
    if "coa" not in st.session_state:
        st.session_state["coa"] = {
            "payem": PAYEM_COA_DEFAULT.copy(),
            "valley": VALLEY_COA_DEFAULT.copy(),
            "ltd": LTD_COA_DEFAULT.copy(),
        }
    if "processed" not in st.session_state:
        st.session_state["processed"] = False


def add_to_history(file_type, keys, file_name):
    h = st.session_state["history"]
    if file_type not in h:
        h[file_type] = {"keys": {}, "count": 0}
    ts = datetime.now().isoformat()
    for k in keys:
        h[file_type]["keys"][k] = {"ts": ts, "file": file_name}
    h[file_type]["count"] = len(h[file_type]["keys"])
    h[file_type]["last_processed"] = ts
    h[file_type]["last_file"] = file_name


def get_coa(coa_type):
    return st.session_state["coa"].get(coa_type, {})


def learn_coa_from_excel(file_bytes):
    """Extract CoA mappings from a corrected Excel file.
    Returns dict with stats and learned mappings."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheets = wb.sheetnames
    result = {"type": None, "new_payem": 0, "new_ltd": 0, "new_valley": 0, "updated": 0, "details": []}

    # Detect type by sheet names
    if "PAYEMDATA" in sheets or any("INC" in s for s in sheets):
        result["type"] = "payem"
        # Read PAYEMDATA sheet (has all transactions with CoA)
        ws = wb["PAYEMDATA"] if "PAYEMDATA" in sheets else None
        if ws:
            headers = [cell.value for cell in ws[1]]
            # Find column indices
            desc_idx = headers.index("תיאור") if "תיאור" in headers else 0
            sub_idx = headers.index("שיוך") if "שיוך" in headers else -1
            coa_cr_idx = headers.index('ח"ן זכות') if 'ח"ן זכות' in headers else -1
            coa_dr_idx = headers.index('ח"ן חובה') if 'ח"ן חובה' in headers else -1

            if coa_cr_idx == -1 or coa_dr_idx == -1:
                result["details"].append('לא נמצאו עמודות ח"ן בגיליון PAYEMDATA')
                return result

            payem_coa = st.session_state["coa"].get("payem", {})
            ltd_coa = st.session_state["coa"].get("ltd", {})

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
                vals = [cell.value for cell in row]
                merchant = str(vals[desc_idx] or "").strip()
                subsidiary = str(vals[sub_idx] or "").strip() if sub_idx >= 0 else ""
                coa_cr = str(vals[coa_cr_idx] or "").strip()
                coa_dr = str(vals[coa_dr_idx] or "").strip()

                if not merchant or not coa_cr or not coa_dr:
                    continue

                # PayEm CoA: key = "merchant|subsidiary"
                payem_key = f"{merchant}|{subsidiary}"
                if payem_key not in payem_coa:
                    payem_coa[payem_key] = [coa_cr, coa_dr]
                    result["new_payem"] += 1
                    result["details"].append(f"PayEm חדש: {payem_key} → [{coa_cr}, {coa_dr}]")
                elif payem_coa[payem_key] != [coa_cr, coa_dr]:
                    payem_coa[payem_key] = [coa_cr, coa_dr]
                    result["updated"] += 1

                # LTD CoA: key = merchant only (for LTD subsidiary)
                if "LTD" in subsidiary:
                    # For LTD, the PAYEMDATA CoA uses 300001 for Inc-side,
                    # but we need the LTD-system CoA (104000 based)
                    # Check if LTD sheet exists for accurate CoA
                    ltd_ws = None
                    for s in sheets:
                        if "LTD" in s and "נגדית" not in s:
                            ltd_ws = wb[s]
                            break

                    if ltd_ws and merchant not in ltd_coa:
                        ltd_headers = [cell.value for cell in ltd_ws[1]]
                        ltd_cr_idx = ltd_headers.index('ח"ן זכות') if 'ח"ן זכות' in ltd_headers else -1
                        ltd_dr_idx = ltd_headers.index('ח"ן חובה') if 'ח"ן חובה' in ltd_headers else -1
                        if ltd_cr_idx >= 0 and ltd_dr_idx >= 0:
                            for ltd_row in ltd_ws.iter_rows(min_row=2, max_row=ltd_ws.max_row, values_only=False):
                                ltd_vals = [cell.value for cell in ltd_row]
                                ltd_merchant = str(ltd_vals[0] or "").strip()
                                if ltd_merchant == merchant:
                                    lcr = str(ltd_vals[ltd_cr_idx] or "").strip()
                                    ldr = str(ltd_vals[ltd_dr_idx] or "").strip()
                                    if lcr and ldr:
                                        ltd_coa[merchant] = [lcr, ldr]
                                        result["new_ltd"] += 1
                                        result["details"].append(f"LTD חדש: {merchant} → [{lcr}, {ldr}]")
                                    break

            st.session_state["coa"]["payem"] = payem_coa
            st.session_state["coa"]["ltd"] = ltd_coa

    elif "VALLEYTRANS" in sheets:
        result["type"] = "valley"
        ws = wb["VALLEYTRANS"]
        headers = [cell.value for cell in ws[1]]
        desc_idx = headers.index("תיאור") if "תיאור" in headers else 1
        cat_idx = headers.index("קטגוריה") if "קטגוריה" in headers else -1
        coa_dr_idx = headers.index('ח"ן חובה') if 'ח"ן חובה' in headers else -1
        coa_cr_idx = headers.index('ח"ן זכות') if 'ח"ן זכות' in headers else -1

        if coa_dr_idx == -1 or coa_cr_idx == -1:
            result["details"].append('לא נמצאו עמודות ח"ן בגיליון VALLEYTRANS')
            return result

        valley_coa = st.session_state["coa"].get("valley", {})

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
            vals = [cell.value for cell in row]
            desc = str(vals[desc_idx] or "").strip()
            category = str(vals[cat_idx] or "").strip() if cat_idx >= 0 else ""
            coa_dr = str(vals[coa_dr_idx] or "").strip()
            coa_cr = str(vals[coa_cr_idx] or "").strip()

            if not desc or not coa_dr or not coa_cr:
                continue

            valley_key = f"{desc}|{category}"
            if valley_key not in valley_coa:
                valley_coa[valley_key] = [coa_dr, coa_cr]
                result["new_valley"] += 1
                result["details"].append(f"Valley חדש: {valley_key} → [{coa_dr}, {coa_cr}]")
            elif valley_coa[valley_key] != [coa_dr, coa_cr]:
                valley_coa[valley_key] = [coa_dr, coa_cr]
                result["updated"] += 1

        st.session_state["coa"]["valley"] = valley_coa

    else:
        result["details"].append("לא זוהה סוג הקובץ - לא נמצאו גיליונות PAYEMDATA או VALLEYTRANS")

    return result


# ===== MAIN APP =====

def main():
    init_session()

    # ===== CUSTOM CSS =====
    st.markdown("""
    <style>
    .main-header {
        background: linear-gradient(135deg, #2e7d32 0%, #66bb6a 100%);
        color: white; padding: 20px; border-radius: 10px; text-align: center; margin-bottom: 20px;
    }
    .main-header h1 { margin: 0; font-size: 1.8em; }
    .main-header p { margin: 5px 0 0; opacity: 0.9; }
    .stat-box {
        background: linear-gradient(135deg, #4caf50, #81c784);
        color: white; padding: 15px; border-radius: 10px; text-align: center;
    }
    .stat-box-blue { background: linear-gradient(135deg, #2196f3, #64b5f6); }
    .stat-box-orange { background: linear-gradient(135deg, #ff9800, #ffb74d); }
    .stat-box-gray { background: linear-gradient(135deg, #78909c, #b0bec5); }
    .badge-valley {
        background: #e8f5e9; color: #2e7d32; border: 2px solid #4caf50;
        padding: 8px 20px; border-radius: 20px; font-weight: bold; display: inline-block;
    }
    .badge-payem {
        background: #e3f2fd; color: #1565c0; border: 2px solid #42a5f5;
        padding: 8px 20px; border-radius: 20px; font-weight: bold; display: inline-block;
    }
    </style>
    """, unsafe_allow_html=True)

    # ===== HEADER =====
    st.markdown("""
    <div class="main-header">
        <h1>ממיר תנועות בנק ואשראי</h1>
        <p>Valley Bank + PayEm → Excel | זיהוי אוטומטי + עיבוד מצטבר</p>
    </div>
    """, unsafe_allow_html=True)

    # ===== SIDEBAR =====
    with st.sidebar:
        st.header("היסטוריה וניהול")
        h = st.session_state["history"]
        v_count = len(h.get("valley", {}).get("keys", {}))
        p_count = len(h.get("payem", {}).get("keys", {}))
        st.metric("Valley Bank תנועות", v_count)
        st.metric("PayEm תנועות", p_count)

        st.divider()
        st.subheader("ייצוא/ייבוא היסטוריה")

        # Export history
        history_json = json.dumps(st.session_state["history"], ensure_ascii=False, indent=2)
        st.download_button("ייצא היסטוריה (JSON)", history_json, "converter_history.json", "application/json")

        # Import history
        hist_file = st.file_uploader("ייבא היסטוריה", type=["json"], key="hist_import")
        if hist_file:
            try:
                imported = json.loads(hist_file.read().decode("utf-8"))
                for t in ["valley", "payem"]:
                    if t in imported and "keys" in imported[t]:
                        if t not in st.session_state["history"]:
                            st.session_state["history"][t] = {"keys": {}, "count": 0}
                        st.session_state["history"][t]["keys"].update(imported[t]["keys"])
                        st.session_state["history"][t]["count"] = len(st.session_state["history"][t]["keys"])
                st.success("ההיסטוריה יובאה בהצלחה!")
            except Exception as e:
                st.error(f"שגיאה בייבוא: {e}")

        # Clear history
        if st.button("נקה היסטוריה", type="secondary"):
            st.session_state["history"] = {"valley": {"keys": {}, "count": 0}, "payem": {"keys": {}, "count": 0}}
            st.success("ההיסטוריה נמחקה!")
            st.rerun()

        st.divider()
        st.subheader('ניהול טבלת ח"ן')

        # Export CoA
        coa_json = json.dumps(st.session_state["coa"], ensure_ascii=False, indent=2)
        st.download_button('ייצא טבלת ח"ן', coa_json, "coa_lookup.json", "application/json")

        # Import CoA from JSON
        coa_file = st.file_uploader('ייבא טבלת ח"ן (JSON)', type=["json"], key="coa_import")
        if coa_file:
            try:
                coa_data = json.loads(coa_file.read().decode("utf-8"))
                st.session_state["coa"] = coa_data
                st.success('טבלת ח"ן עודכנה בהצלחה!')
            except Exception as e:
                st.error(f"שגיאה: {e}")

        # Learn CoA from corrected Excel
        st.divider()
        st.subheader('למידת ח"ן מאקסל מתוקן')
        st.caption('העלה אקסל מתוקן כדי ללמוד חשבונות חדשים')
        learn_file = st.file_uploader('העלה אקסל מתוקן', type=["xlsx"], key="coa_learn")
        if learn_file:
            try:
                result = learn_coa_from_excel(learn_file.read())
                if result["type"] is None:
                    st.warning("לא זוהה סוג הקובץ")
                else:
                    total_new = result["new_payem"] + result["new_ltd"] + result["new_valley"]
                    total_changes = total_new + result["updated"]
                    if total_changes == 0:
                        st.info('לא נמצאו חשבונות חדשים - טבלת ח"ן כבר מעודכנת')
                    else:
                        if result["new_payem"] > 0:
                            st.success(f'{result["new_payem"]} חשבונות PayEm חדשים')
                        if result["new_ltd"] > 0:
                            st.success(f'{result["new_ltd"]} חשבונות LTD חדשים')
                        if result["new_valley"] > 0:
                            st.success(f'{result["new_valley"]} חשבונות Valley חדשים')
                        if result["updated"] > 0:
                            st.info(f'{result["updated"]} חשבונות עודכנו')
                        with st.expander("פרטים"):
                            for d in result["details"][:50]:
                                st.text(d)
            except Exception as e:
                st.error(f"שגיאה בקריאת הקובץ: {e}")

    # ===== FILE UPLOAD =====
    uploaded_file = st.file_uploader("העלה קובץ CSV (Valley Bank / PayEm)", type=["csv"], key="csv_upload")

    if uploaded_file is not None:
        # Parse CSV
        content = uploaded_file.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        # Detect file type
        file_type = detect_file_type(rows)

        if file_type == "valley":
            st.markdown('<div style="text-align:center"><span class="badge-valley">Valley Bank</span></div>', unsafe_allow_html=True)
        elif file_type == "payem":
            st.markdown('<div style="text-align:center"><span class="badge-payem">PayEm Card</span></div>', unsafe_allow_html=True)
        else:
            st.error("לא ניתן לזהות את סוג הקובץ")
            return

        st.info(f"קובץ: **{uploaded_file.name}** | גודל: {uploaded_file.size / 1024:.1f} KB")

        # Parse transactions
        if file_type == "valley":
            all_parsed = parse_valley(rows, get_coa("valley"), VALLEY_CAT_COA)
        else:
            all_parsed = parse_payem(rows, get_coa("payem"), get_coa("ltd"))

        # Dedup
        new_txns, dupes = analyze_duplicates(file_type, all_parsed, st.session_state["history"])

        # Show stats
        col1, col2, col3 = st.columns(3)
        col1.metric("סה״כ תנועות", len(all_parsed))
        col2.metric("חדשות", len(new_txns))
        col3.metric("כבר עובדו", len(dupes))

        # Process options
        if len(new_txns) == 0:
            # ALL transactions already processed - show re-download option
            st.warning("כל התנועות כבר עובדו בעבר")
            redownload = st.checkbox("הורד מחדש את כל התנועות (למקרה שההורדה הקודמת נכשלה)", value=False)
            if not redownload:
                return
            data = all_parsed
        elif len(dupes) > 0:
            new_only = st.checkbox("עבד רק תנועות חדשות (דלג על כפילויות)", value=True)
            data = new_txns if new_only else all_parsed
        else:
            data = all_parsed

        # ===== PREVIEW TABLE =====
        st.subheader("תצוגה מקדימה")

        if file_type == "valley":
            import pandas as pd
            preview_data = []
            for t in data[:50]:
                preview_data.append({
                    "תאריך": t["date_formatted"],
                    "תיאור": t["description"],
                    "חובה": t["amount"],
                    "זכות": t["amount"],
                    "קטגוריה": t["category"],
                    'ח"ן חובה': t["coa_debit"],
                    'ח"ן זכות': t["coa_credit"],
                })
            df = pd.DataFrame(preview_data)
            st.dataframe(df, use_container_width=True, hide_index=True)

            # Stats
            total_amount = sum(t["amount"] for t in data)
            no_coa = sum(1 for t in data if not t["coa_debit"])
            scol1, scol2, scol3 = st.columns(3)
            scol1.metric("תנועות", len(data))
            scol2.metric("סך תנועות", f"${total_amount:,.2f}")
            if no_coa > 0:
                scol3.metric('ללא ח"ן', no_coa)

        else:
            import pandas as pd
            preview_data = []
            for t in data[:50]:
                preview_data.append({
                    "תיאור": t["description"],
                    "סכום": t["net_amount"],
                    "זכות": t["credit_amount"] if t["credit_amount"] is not None else "",
                    "חובה": t["debit_amount"] if t["debit_amount"] is not None else "",
                    "אסמכתא 1": t["ref1"],
                    "אסמכתא 2": t["ref2"],
                    'ח"ן זכות': t["coa_credit"],
                    'ח"ן חובה': t["coa_debit"],
                    "תאריך": t["date_formatted"],
                    "שיוך": t["subsidiary"],
                })
            df = pd.DataFrame(preview_data)
            st.dataframe(df, use_container_width=True, hide_index=True)

            # Stats
            total_debit = sum(t["debit_amount"] or 0 for t in data)
            total_credit = sum(t["credit_amount"] or 0 for t in data)
            no_coa = sum(1 for t in data if not t["coa_credit"])
            scol1, scol2, scol3, scol4 = st.columns(4)
            scol1.metric("תנועות", len(data))
            scol2.metric("סך חובה", f"${total_debit:,.2f}")
            scol3.metric("סך זכות", f"${total_credit:,.2f}")
            if no_coa > 0:
                scol4.metric('ללא ח"ן', no_coa)

        st.markdown(f"מציג {min(50, len(data))} מתוך {len(data)} תנועות")

        # ===== EXCEL DOWNLOAD =====
        st.divider()

        if file_type == "valley":
            wb = build_valley_excel(data)
            checks = None
        else:
            wb, checks = build_payem_excel(data)

        today = datetime.now()
        prefix = "Valley_Bank" if file_type == "valley" else "PayEm"
        filename = f"{prefix}_{today.day}_{today.month}_{today.year}.xlsx"

        # Show integrity checks for PayEm
        if checks:
            st.subheader("בדיקות שלמות")
            c1, c2, c3 = st.columns(3)
            with c1:
                if checks["count_ok"]:
                    st.success(f"שורות: DATA({checks['count_data']}) = INC({checks['count_inc']}) + LTD({checks['count_ltd']})")
                else:
                    st.error(f"שורות: DATA({checks['count_data']}) != INC({checks['count_inc']}) + LTD({checks['count_ltd']})")
            with c2:
                if checks["debit_ok"]:
                    st.success(f"חובה: ${checks['data_debit']:,.2f} = INC(${checks['inc_debit']:,.2f}) + LTD(${checks['ltd_debit']:,.2f})")
                else:
                    st.error(f"חובה: DATA ${checks['data_debit']:,.2f} != INC(${checks['inc_debit']:,.2f}) + LTD(${checks['ltd_debit']:,.2f})")
            with c3:
                if checks["credit_ok"]:
                    st.success(f"זכות: ${checks['data_credit']:,.2f} = INC(${checks['inc_credit']:,.2f}) + LTD(${checks['ltd_credit']:,.2f})")
                else:
                    st.error(f"זכות: DATA ${checks['data_credit']:,.2f} != INC(${checks['inc_credit']:,.2f}) + LTD(${checks['ltd_credit']:,.2f})")

        excel_bytes = workbook_to_bytes(wb)

        # Check if these transactions are already saved
        saved_keys = st.session_state["history"].get(file_type, {}).get("keys", {})
        data_keys = {t["key"] for t in data}
        already_saved = data_keys.issubset(set(saved_keys.keys()))

        # Step 1: Download Excel
        st.download_button(
            f"📥 הורד קובץ Excel ({len(data)} תנועות)",
            excel_bytes,
            filename,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )

        # Step 2: Confirm save to history (separate action)
        if already_saved:
            st.success("✅ תנועות אלו כבר שמורות בהיסטוריה")
        else:
            st.info("⚠️ לאחר שווידאת שהקובץ הורד בהצלחה - לחץ לשמור בהיסטוריה:")
            if st.button("💾 הורדתי בהצלחה - שמור בהיסטוריה", type="secondary"):
                keys = [t["key"] for t in data]
                add_to_history(file_type, keys, uploaded_file.name)
                st.success(f"✅ {len(data)} תנועות נשמרו בהיסטוריה!")
                st.rerun()


if __name__ == "__main__":
    main()
